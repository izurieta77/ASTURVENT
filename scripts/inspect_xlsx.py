"""Print high-level structure of every xlsx in data/raw/.

For each workbook:
  - list sheet names with their dimensions
  - print the first 5 rows of each sheet
"""
import openpyxl
import os
from pathlib import Path

RAW = Path("data/raw")

def trim(v, n=30):
    if v is None:
        return ""
    s = str(v)
    return s if len(s) <= n else s[:n] + "..."

for xlsx in sorted(RAW.glob("*.xlsx")):
    print("="*100)
    print(f"FILE: {xlsx.name}")
    print("="*100)
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n  SHEET: {sn!r}  dims={ws.max_row}r x {ws.max_column}c")
        try:
            it = ws.iter_rows(min_row=1, max_row=5, values_only=True)
            for i, row in enumerate(it, 1):
                cells = [trim(c) for c in row[:15]]
                print(f"    r{i}: {cells}")
        except Exception as e:
            print(f"    (err: {e})")
    wb.close()
