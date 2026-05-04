"""Consolidate per-client spreadsheets into a single master xlsx.

Reads every workbook in data/raw/, scans each sheet, autodetects header rows,
maps source columns to a unified schema, and writes one tab per client to
data/master/CONSOLIDADO_MAESTRO_Clientes.xlsx.
"""
from __future__ import annotations

import datetime as dt
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RAW = Path("data/raw")
OUT = Path("data/master/CONSOLIDADO_MAESTRO_Clientes.xlsx")

CLIENTS = [
    ("01_lavanderia_manantiales.xlsx", "LAVANDERIA LOS MANANTIALES"),
    ("02_transportes_dehui.xlsx", "TRANSPORTES DEHUI"),
    ("03_don_harina_wawita.xlsx", "DON HARINA - WAWITA"),
    ("04_thd_gio.xlsx", "THD GIO"),
    ("05_reca_comunicaciones.xlsx", "RECA COMUNICACIONES"),
]

# Sheets to skip:
#   - catalogs (CATÁLOGOS) and audit reports (_AUDIT_*) are not transactions
#   - Historico_Backup / Pagos_Saldos / CORRECTA SALDOS / Copia de Hoja 1 are
#     known to mirror data already present in the primary sheets; including them
#     just inflates the file with rows that fail dedup due to minor format diffs
SKIP_SHEETS = {
    "CATÁLOGOS", "CATALOGOS",
    "_AUDIT_Consolidacion", "_audit_consolidacion",
    "Sheet2", "Hoja 2",
    "Historico_Backup",
    "Pagos_Saldos",
    "CORRECTA SALDOS",
    "Copia de Hoja 1",
}

# Empty columns (Codigo_Autz, Precintos, saldos) are dropped — none of the
# clients have any data there. Foto_*, Tipo_Documento, Precio_Unit, Odometro,
# Firma_Chofer have <2% fill but are kept for completeness as the user asked
# for foto links and "extras".
UNIFIED_COLS = [
    "Fecha", "Hora", "Ticket", "Cantidad", "Producto", "Monto", "Folio",
    "Chofer", "Placas",
    "Tipo_Documento", "Precio_Unit", "Odometro",
    "Foto_Placa", "Foto_Ticket", "Firma_Chofer",
    "Fuente",
]

# Short codes for Fuente_Hoja to save bytes on 30k-row sheets.
FUENTE_SHORT = {
    "Despachos_SGM": "SGM",
    "Despachos_SGM_APP": "APP",
    "Hoja 1": "H1",
    "Sheet1": "H1",
    "TORRE_DE_CONTROL": "TDC",
    "Reconstruido": "RCN",
}

# Map source header (uppercased, stripped) -> unified column
HEADER_ALIASES = {
    # Fecha
    "FECHA": "Fecha",
    # Hora
    "HORA": "Hora", "HR": "Hora",
    # Ticket
    "TICKET": "Ticket", "NUMERO DE TICKET": "Ticket", "NÚMERO DE TICKET": "Ticket",
    "TICKET_EGAS": "Ticket", "N° DE TICKET": "Ticket", "NO. DE TICKET": "Ticket",
    # Cantidad
    "CANTIDAD": "Cantidad", "LITROS": "Cantidad", "LTS": "Cantidad",
    # Producto
    "PRODUCTO": "Producto", "DIÉSEL UBA": "Producto", "DIESEL UBA": "Producto",
    # Monto
    "MONTO": "Monto", "IMPORTE": "Monto", "TOTAL": "Monto",
    # Folio
    "FOLIO_VALE": "Folio", "FOLIO": "Folio", "VALE": "Folio",
    # Chofer / placas
    "CHOFER": "Chofer", "OPERADOR": "Chofer",
    "PLACAS": "Placas", "PLACA": "Placas",
    # Doc app
    "PRECIO_UNIT": "Precio_Unit", "PRECIO UNITARIO": "Precio_Unit",
    "ODOMETRO": "Odometro", "ODÓMETRO": "Odometro",
    "CODIGO_AUTZ": "Codigo_Autz", "CÓDIGO_AUTZ": "Codigo_Autz",
    "TIPO_DOCUMENTO": "Tipo_Documento",
    "PREFIX": "Prefix",
    # Fotos / firma
    "FOTO_PLACA": "Foto_Placa", "FOTO PLACA": "Foto_Placa",
    "FOTO_TICKET": "Foto_Ticket", "FOTO TICKET": "Foto_Ticket",
    "FIRMA_CHOFER": "Firma_Chofer", "FIRMA CHOFER": "Firma_Chofer",
    # Saldos
    "ABONO": "Abono",
    "SUMA CONSUMOS": "Suma_Consumos", "SUMA_CONSUMOS": "Suma_Consumos",
    "SALDO": "Saldo",
    "SALDO ANTERIOR": "Saldo_Anterior", "SALDO_ANTERIOR": "Saldo_Anterior",
    # Precintos
    "PRECINTO_1_ENTRADA": "Precinto_1_Entrada", "PRECINTO 1 ENTRADA": "Precinto_1_Entrada",
    "PRECINTO_2_ENTRADA": "Precinto_2_Entrada", "PRECINTO 2 ENTRADA": "Precinto_2_Entrada",
    "PRECINTO_1_SALIDA": "Precinto_1_Salida", "PRECINTO 1 SALIDA": "Precinto_1_Salida",
    "PRECINTO_2_SALIDA": "Precinto_2_Salida", "PRECINTO 2 SALIDA": "Precinto_2_Salida",
    # Reconstruido (THD GIO)
    "FUENTE": "Fuente",
}

# Tokens that strongly suggest a row is a header row.
HEADER_TOKENS = {"FECHA", "TICKET", "MONTO", "PRODUCTO", "CHOFER", "CANTIDAD", "LITROS", "FOLIO"}


def normalise(v):
    """Return v rendered for output: dates as YYYY-MM-DD, datetimes split, others stripped."""
    if v is None:
        return ""
    if isinstance(v, dt.datetime):
        # Some sheets put time as a datetime starting at midnight; render only the date.
        if v.hour == 0 and v.minute == 0 and v.second == 0:
            return v.strftime("%Y-%m-%d")
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, dt.date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, dt.time):
        return v.strftime("%H:%M:%S")
    if isinstance(v, float):
        # Tickets / folios often come as 638953.0 -> 638953
        if v.is_integer():
            return str(int(v))
        return str(v)
    return str(v).strip()


def header_score(row_vals):
    """Return how many HEADER_TOKENS appear in this row (uppercased)."""
    score = 0
    for v in row_vals:
        if v is None:
            continue
        s = str(v).strip().upper()
        if s in HEADER_TOKENS:
            score += 1
    return score


def find_header(rows):
    """Find the best header row in the first ~5 rows. Returns (idx, mapping)
    where mapping is {col_index: unified_col_name}. Returns (None, None) if no
    plausible header was found."""
    best_idx, best_score, best_map = None, 0, None
    for idx, row in enumerate(rows[:6]):
        s = header_score(row)
        if s < 2:
            continue
        m = {}
        for ci, v in enumerate(row):
            if v is None:
                continue
            key = str(v).strip().upper()
            if key in HEADER_ALIASES:
                m[ci] = HEADER_ALIASES[key]
        if s > best_score and m:
            best_idx, best_score, best_map = idx, s, m
    return best_idx, best_map


def is_blank(row):
    return all(v in (None, "") for v in row)


def has_any_key_field(unified_row):
    """Row counts as informative if it has at least one of the 7 key fields."""
    keys = ("Fecha", "Hora", "Ticket", "Cantidad", "Producto", "Monto", "Folio")
    return any(unified_row.get(k) not in (None, "", "-") for k in keys)


def process_workbook(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out_rows = []
    stats = {}
    for sn in wb.sheetnames:
        if sn.strip() in SKIP_SHEETS:
            stats[sn] = "skip"
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            stats[sn] = "empty"
            continue
        head_idx, mapping = find_header(rows)
        if mapping is None:
            stats[sn] = "no-header"
            continue
        kept = 0
        for row in rows[head_idx + 1:]:
            if is_blank(row):
                continue
            unified = {c: "" for c in UNIFIED_COLS}
            for ci, ucol in mapping.items():
                if ci < len(row):
                    unified[ucol] = normalise(row[ci])
            if not has_any_key_field(unified):
                continue
            unified["Fuente"] = FUENTE_SHORT.get(sn, sn)
            out_rows.append(unified)
            kept += 1
        stats[sn] = f"kept {kept} (header r{head_idx+1})"
    wb.close()
    return out_rows, stats


def dedup_key(row):
    """A despacho is uniquely identified by date+ticket+cantidad. Using monto
    or hora here causes false negatives because formats vary across sheets."""
    return (
        row.get("Fecha", ""),
        row.get("Ticket", ""),
        row.get("Cantidad", ""),
    )


def dedup(rows):
    seen = {}
    for r in rows:
        k = dedup_key(r)
        # Keep the row that has the most non-empty fields (richer source).
        if k not in seen:
            seen[k] = r
        else:
            cur = seen[k]
            if sum(1 for v in r.values() if v) > sum(1 for v in cur.values() if v):
                seen[k] = r
    return list(seen.values())


def sort_key(r):
    f = r.get("Fecha", "")
    # Try to parse as ISO date; fall back to original string for ordering.
    try:
        return (0, dt.datetime.strptime(f[:10], "%Y-%m-%d"), r.get("Hora", ""))
    except Exception:
        return (1, f, r.get("Hora", ""))


def build_master(per_client: dict):
    OUT.parent.mkdir(parents=True, exist_ok=True)
    out = openpyxl.Workbook()
    out.remove(out.active)

    summary = out.create_sheet("RESUMEN")
    summary.append(["Cliente", "Filas consolidadas", "Pestañas fuente"])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for client_name, (rows, stats) in per_client.items():
        ws = out.create_sheet(client_name[:31])  # excel sheet name limit
        ws.append(UNIFIED_COLS)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for r in rows:
            ws.append([r.get(c, "") for c in UNIFIED_COLS])
        ws.freeze_panes = "A2"
        # Reasonable column widths.
        widths = {
            "Fecha": 12, "Hora": 10, "Ticket": 12, "Cantidad": 10,
            "Producto": 12, "Monto": 11, "Folio": 18,
            "Chofer": 24, "Placas": 12,
            "Tipo_Documento": 12, "Precio_Unit": 11, "Odometro": 11,
            "Foto_Placa": 11, "Foto_Ticket": 11, "Firma_Chofer": 11,
            "Fuente": 8,
        }
        for i, c in enumerate(UNIFIED_COLS, 1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 14)

        sources = ", ".join(f"{k}: {v}" for k, v in stats.items())
        summary.append([client_name, len(rows), sources])

    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 22
    summary.column_dimensions["C"].width = 120
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
    summary.freeze_panes = "A2"

    out.save(OUT)
    print(f"\nSaved: {OUT}")


def main():
    per_client = {}
    for fname, client in CLIENTS:
        path = RAW / fname
        rows, stats = process_workbook(path)
        before = len(rows)
        rows = dedup(rows)
        after = len(rows)
        rows.sort(key=sort_key)
        per_client[client] = (rows, stats)
        print(f"{client}: rows {before} -> {after} after dedup")
        for sn, st in stats.items():
            print(f"  - {sn}: {st}")

    build_master(per_client)


if __name__ == "__main__":
    main()
